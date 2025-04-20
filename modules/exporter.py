# modules/exporter.py
"""Exporta la tabla de amortización y los indicadores a:
    1. Un archivo Excel con estructura tabular, estilos y un apartado de decisiones.
    2. Un CSV con la tabla de amortización lisa (sin estilo) para ingestión rápida.

Pensado para ser la única dependencia de salida tabular.  Más adelante podremos
inyectar aquí funciones para embellecer el archivo (formatos condicionales,
colores corporativos, fórmulas avanzadas, etc.).  Por ahora la prioridad es
una base sólida y editable.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment

# ===========================
# Configuración básica global
# ===========================

# Estilo por defecto para las tablas Excel (puedes cambiarlo después).
_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
)

# Reglas simplificadas para clasificar escenarios
# (ajusta los umbrales a tu criterio empresarial)
_DEFAULT_RULES = {
    "favorable": lambda ind: ind["VPN ($)"] > 0 and ind["TIR (%)"] >= 0 and ind["Periodo de Recuperación (meses)"] is not None and ind["Periodo de Recuperación (meses)"] <= 60,
    "poco_favorable": lambda ind: ind["VPN ($)"] <= 0 and ind["TIR (%)"] >= 0,
    # todo: afinar criterios intermedios
}


# =========================
# Funciones públicas del módulo
# =========================

def exportar_excel(
    ruta_salida: str | Path,
    df_amort: pd.DataFrame,
    indicadores: Dict[str, float | int | None],
    nombre_escenario: str,
    reglas: Optional[Dict[str, callable]] = None,
) -> None:
    """Genera un archivo Excel con tres hojas:

    Amortización   – Tabla estructurada con estilo.
    Indicadores    – Indicadores clave + campos reservados.
    Decisiones     – Clasificación del escenario y glosario.
    """

    reglas = reglas or _DEFAULT_RULES

    # --------------------------
    # Crear libro / hoja 1 (amortización)
    # --------------------------
    wb = Workbook()
    ws_amort = wb.active
    ws_amort.title = "Amortización"

    for row in dataframe_to_rows(df_amort, index=False, header=True):
        ws_amort.append(row)

    # Aplicar encabezado en negrita
    for cell in ws_amort[1]:
        cell.font = Font(bold=True)

    # Definir la tabla formal (para filtros y estilos nativos de Excel)
    ultima_fila = ws_amort.max_row
    ultima_col = ws_amort.max_column
    tabla = Table(ref=f"A1:{chr(64+ultima_col)}{ultima_fila}", displayName="TablaAmortizacion", tableStyleInfo=_TABLE_STYLE)
    ws_amort.add_table(tabla)

    # Columnas con ancho auto-ajustado básico
    for col in ws_amort.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None)
        ws_amort.column_dimensions[col[0].column_letter].width = max_len + 2

    # --------------------------
    # Hoja 2 – Indicadores
    # --------------------------
    ws_ind = wb.create_sheet("Indicadores")
    ws_ind["A1"] = "Escenario:"
    ws_ind["B1"] = nombre_escenario
    ws_ind["A1"].font = Font(bold=True)

    row = 3
    for k, v in indicadores.items():
        ws_ind[f"A{row}"] = k
        ws_ind[f"A{row}"].font = Font(bold=True)
        ws_ind[f"B{row}"] = v
        row += 1

    # Espacio reservado para fórmulas adicionales (punto de equilibrio, etc.)
    ws_ind[f"A{row+1}"] = "Punto de Equilibrio ($)"
    ws_ind[f"A{row+1}"].font = Font(bold=True)
    ws_ind[f"B{row+1}"] = ""  # la fórmula se añadirá más adelante por el usuario

    # --------------------------
    # Hoja 3 – Decisiones
    # --------------------------
    ws_dec = wb.create_sheet("Decisiones")

    cabecera = [
        "Escenario",
        "TIR (%)",
        "VPN ($)",
        "Periodo de Recuperación (meses)",
        "Clasificación",
    ]
    ws_dec.append(cabecera)
    for cell in ws_dec[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Clasificación según reglas
    clasificacion = _clasificar(indicadores, reglas)

    ws_dec.append([
        nombre_escenario,
        indicadores.get("TIR (%)"),
        indicadores.get("VPN ($)"),
        indicadores.get("Periodo de Recuperación (meses)"),
        clasificacion.title().replace("_", " "),
    ])

    # Ancho de columnas
    for col in ws_dec.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None)
        ws_dec.column_dimensions[col[0].column_letter].width = max_len + 2

    # Pequeño glosario / guía de interpretación debajo
    start_row = ws_dec.max_row + 3
    ws_dec[f"A{start_row}"] = "Guía de interpretación:"
    ws_dec[f"A{start_row}"].font = Font(bold=True)

    _insert_glosario(ws_dec, start_row + 1)

    # --------------------------
    # Salvar workbook
    # --------------------------
    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)


def exportar_csv(ruta_csv: str | Path, df_amort: pd.DataFrame) -> None:
    """Exporta la tabla de amortización a CSV (delimitador ;, encoding utf‑8)."""
    ruta_csv = Path(ruta_csv)
    ruta_csv.parent.mkdir(parents=True, exist_ok=True)
    df_amort.to_csv(ruta_csv, index=False, sep=";", encoding="utf-8")


# =========================
# Ayudantes privados
# =========================

def _clasificar(indicadores: Dict[str, float | int | None], reglas) -> str:
    """Devuelve la clave de la regla que se cumpla primero.  Si ninguna,
    regresa 'intermedio'."""
    for nombre, fn in reglas.items():
        try:
            if fn(indicadores):
                return nombre
        except Exception:
            # Si los datos son insuficientes para evaluar, continuamos
            continue
    return "intermedio"


def _insert_glosario(ws, start_row: int):
    """Inserta texto descriptivo en la hoja de decisiones."""
    notas = [
        ("Favorable", "VPN positivo, TIR por encima del costo de capital, corto periodo de recuperación."),
        ("Intermedio", "Mixto: uno o dos indicadores aceptables pero con reservas. Requiere análisis cualitativo."),
        ("Poco favorable", "VPN negativo y/o TIR baja, periodo de recuperación muy largo.")
    ]
    for i, (titulo, descripcion) in enumerate(notas, start=0):
        fila = start_row + i
        ws[f"A{fila}"] = titulo
        ws[f"A{fila}"].font = Font(bold=True)
        ws[f"B{fila}"] = descripcion
        ws[f"B{fila}"].alignment = Alignment(wrap_text=True)


# ==================================
# Ejemplo de uso rápido (solo tests)
# ==================================
if __name__ == "__main__":  # pragma: no cover
    # Carga de un ejemplo pequeño si se ejecuta directamente.
    datos_demo = pd.DataFrame(
        {
            "Mes": [1, 2, 3],
            "Cuota ($)": [1000, 1000, 1000],
            "Interés ($)": [500, 480, 460],
            "Amortización ($)": [500, 520, 540],
            "Saldo ($)": [9500, 8980, 8440],
            "Seguro ($)": [30, 30, 30],
            "Flujo ($)": [-1030, -1030, -1030],
        }
    )
    indicadores_demo = {
        "TIR (%)": 12.5,
        "VPN ($)": 15000,
        "Periodo de Recuperación (meses)": 48,
    }
    exportar_excel("demo.xlsx", datos_demo, indicadores_demo, "Escenario Demo")
    exportar_csv("demo.csv", datos_demo)
